#!/usr/bin/env python3
"""
由 JSON spec 生成舆情检索任务清单 Excel（三 sheet）。

用法:
    python build_task_xlsx.py <spec.json> <输出.xlsx>

spec 结构见 assets/task_spec_example.json。
"""
import json
import sys
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- 样式 ----------
HFONT = Font(name="Noto Sans CJK SC", size=10, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F3864")
DFONT = Font(name="Arial", size=10)
MONO = Font(name="Consolas", size=9)
_thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="top", wrap_text=True)
TITLE_FONT = Font(name="Noto Sans CJK SC", size=13, bold=True, color="1F3864")

# 频次色标：越高频越醒目
FREQ_FILL = {
    "快讯/小时级": PatternFill("solid", fgColor="FF851B"),
    "日级": PatternFill("solid", fgColor="FF4136"),
    "周级": PatternFill("solid", fgColor="FFDC00"),
    "双周级": PatternFill("solid", fgColor="DDDDDD"),
    "月级": PatternFill("solid", fgColor="EEEEEE"),
}
FREQ_FONT = {
    "快讯/小时级": Font(name="Arial", size=10, bold=True, color="FFFFFF"),
    "日级": Font(name="Arial", size=10, bold=True, color="FFFFFF"),
}
TASK_HEADERS = [
    "任务ID", "检索组", "国家/地区", "语种", "检索式(布尔)", "检索式(Google语法)",
    "目标信源白名单(域名)", "建议频次", "命中期望相关度", "状态", "运营注/说明",
]
TASK_WIDTHS = [9, 20, 14, 8, 60, 50, 46, 14, 14, 8, 50]

KW_HEADERS = ["层", "键类别", "关键词/别名", "语种", "context_guard(短缩写强制AND)", "排除词/备注"]
KW_WIDTHS = [6, 20, 55, 10, 50, 40]

DEFAULT_NOTES = [
    ("用途", "每一行=一个爬取任务。调度器按「任务ID」遍历，用「检索式」构造查询，限定「目标信源白名单」域名，"
             "按「建议频次」排程。运营可在「状态」列勾选 待启用/运行中/暂停"),
    ("频次色标", "红=日级 橙=快讯小时级 黄=周级 灰=双周/月级。涉人员安全的组建议独立快讯管道，命中即推送，不等轮询"),
    ("双轨三式结构", "全量新闻轨=实体键宽召回；负面新闻轨=实体键 AND 风险词，只抓负面，可高频；"
                  "行业新闻轨=地名键 AND 行业词 AND 外资标识，抓同区域未点名负面，命中多落在 context 层"),
    ("语法差异", "布尔式用于新闻API/舆情库，Google式用于搜索引擎抓取。取反写法不同(NOT vs 逐词-)，"
              "嵌套深度不同，不可混用，混用会静默失效"),
    ("去重", "企业通稿多站转载，抓取后须 URL 归一化 + 正文相似度去重，否则正面量级虚高、指标失真"),
    ("本国噪音隔离", "母语检索式必须附加地域词或按域名分流，否则国内招采、人事、培训信息会稀释海外信号"),
    ("单一信源标注", "只有一个来源的事件，引用前必须交叉验证；自述由 AI 撰写的站点禁作独立依据，只能当线索"),
    ("时间窗", "常规统计窗为最近12个月。更早报道可用于确认项目背景与历史先例，但不计入当期统计"),
    ("未验证检索式", "机器生成但未做召回验证的检索式(尤其小语种)须标注待验证，先设低频，人工抽查后再提频"),
    ("相关度口径", "direct=监测主体自身/在场承包；indirect=关联实体或业主(舆情传导)；context=同区域行业性报道。"
                "统计客户负面时分层出数，三层不可直接合计"),
]


def _write_header(ws, headers, widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font, cell.fill, cell.alignment, cell.border = HFONT, HFILL, CTR, BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_task_sheet(ws, tasks):
    _write_header(ws, TASK_HEADERS, TASK_WIDTHS)
    for ri, t in enumerate(tasks, 2):
        vals = [
            t.get("id", ""), t.get("group", ""), t.get("region", ""), t.get("lang", ""),
            t.get("boolean", ""), t.get("google", ""),
            "; ".join(t.get("sources", [])) if isinstance(t.get("sources"), list) else t.get("sources", ""),
            t.get("frequency", ""), t.get("relevance", ""),
            t.get("status", "待启用"), t.get("note", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border = BORDER
            cell.font = MONO if ci in (5, 6, 7) else DFONT
            cell.alignment = WRAP if ci in (5, 6, 7, 11) else CTR
        freq = t.get("frequency", "")
        if freq in FREQ_FILL:
            ws.cell(ri, 8).fill = FREQ_FILL[freq]
        if freq in FREQ_FONT:
            ws.cell(ri, 8).font = FREQ_FONT[freq]
    ws.auto_filter.ref = f"A1:K{len(tasks) + 1}"


def build_kw_sheet(ws, keywords):
    _write_header(ws, KW_HEADERS, KW_WIDTHS)
    for ri, k in enumerate(keywords, 2):
        vals = [
            k.get("layer", ""), k.get("category", ""), k.get("terms", ""),
            k.get("lang", ""), k.get("guard", ""), k.get("note", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border = BORDER
            cell.font = MONO if ci in (3, 5) else DFONT
            cell.alignment = WRAP if ci in (3, 5, 6) else CTR
    ws.auto_filter.ref = f"A1:F{len(keywords) + 1}"


def build_notes_sheet(ws, title, extra_notes):
    rows = [(title, "")] + [("", "")] + DEFAULT_NOTES + list(extra_notes)
    for ri, (a, b) in enumerate(rows, 1):
        ws.cell(ri, 1, a)
        ws.cell(ri, 2, b)
        ws.cell(ri, 1).font = Font(name="Arial", size=10, bold=not b)
        ws.cell(ri, 2).font = Font(name="Arial", size=10)
        ws.cell(ri, 1).alignment = WRAP
        ws.cell(ri, 2).alignment = WRAP
    ws.cell(1, 1).font = TITLE_FONT
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 100


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    spec_path, out_path = sys.argv[1], sys.argv[2]
    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    tasks = spec.get("tasks", [])
    keywords = spec.get("keywords", [])
    if not tasks:
        print("警告: spec 中没有 tasks，生成的清单将为空")

    wb = openpyxl.Workbook()
    build_task_sheet(wb.active, tasks)
    wb.active.title = "检索任务清单"
    build_kw_sheet(wb.create_sheet("关键词字典"), keywords)
    build_notes_sheet(
        wb.create_sheet("调度说明"),
        spec.get("title", "舆情爬虫检索任务清单 · 使用说明"),
        [(n.get("key", ""), n.get("value", "")) for n in spec.get("extra_notes", [])],
    )

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print(f"已生成: {out_path}")
    print(f"  检索任务: {len(tasks)} 条 | 关键词字典: {len(keywords)} 行")
    freqs = {}
    for t in tasks:
        freqs[t.get("frequency", "未设")] = freqs.get(t.get("frequency", "未设"), 0) + 1
    print("  频次分布: " + " | ".join(f"{k}:{v}" for k, v in freqs.items()))


if __name__ == "__main__":
    main()
